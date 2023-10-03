# imports
import math
import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import openpyxl 

A1 = 0.51

w0 = 2*np.pi*1e3
k = 15.85
B = 2*np.pi*250

def T_dB(x, k, w0, B):
    return 20 * np.log10(np.abs(k*B*1j*x / (w0**2 + 1j * x * B - x**2)))

# Caminho relativo para o arquivo Excel dentro da pasta "lab_04"
caminho_arquivo_excel = 'Lab_04/2_6_3.xlsx'

# Carregando o arquivo Excel
workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
sheet1 = workbook['dados_fase']
sheet2 = workbook['dados_ganho']

freqs, A2, phases = [], [], []

for row in sheet1.iter_rows(min_row=2, values_only=True):
    f, phase = row[0:2]
    freqs.append(float(f))
    phases.append(float(phase))

for row in sheet2.iter_rows(min_row=2, values_only=True):
    A = row[1]
    A2.append(float(A))

w = [2*np.pi * f * 1e3 for f in freqs]
A2_dB = [20*np.log10(np.abs(a/A1)) for a in A2]

#  curva teórica
w_values = np.linspace(0, max(w), 10000)

A2_magnitude_dB = T_dB(w_values, k, w0, B)

A2_initial_guess = (k, w0, B)  # (K, w0, B)

A2_params, A2_covariance = curve_fit(T_dB, w, A2_dB, p0=A2_initial_guess)
A2_x_fit = np.linspace(min(w), max(w), 1000)
A2_y_fit = T_dB(A2_x_fit, *A2_params)

plt.plot(w_values, A2_magnitude_dB, label='Curva Teórica', color='green')
plt.scatter(w, A2_dB, label = "Pontos experimentais")
plt.plot(A2_x_fit, np.real(A2_y_fit), 'g-', label='Ajuste', color='red')
plt.xlim(2000, 35000)
plt.ylim(0, 30)
plt.xscale('log')
plt.xlabel('w [rad/s]')
plt.ylabel('|T| [dB]')
plt.legend()
plt.grid()
plt.show()

print("Parâmetros Ajustados para o ajuste ao ganho em dB:")
print("K:", A2_params[0])
print("w0:", A2_params[1])
print("B:", A2_params[2])

phases_new = [p if p>0 else p+360 for p in phases]

def phase_T_dB(x, k, w0, B):
    return np.degrees(np.angle(k*B*1j*x / (w0**2 + 1j * x * B - x**2))) + 180

A2_phase = phase_T_dB(w_values, k, w0, B)

A2_params_phase, A2_covariance_phase = curve_fit(phase_T_dB, w, phases_new, p0=A2_initial_guess)
A2_x_fit_phase = np.linspace(min(w), max(w), 1000)
A2_y_fit_phase = phase_T_dB(A2_x_fit_phase, *A2_params_phase)

plt.plot(w_values, A2_phase, label='Curva Teórica', color='green')
plt.scatter(w, phases_new, label = "Pontos experimentais")
plt.plot(A2_x_fit_phase, np.real(A2_y_fit_phase), 'g-', label='Ajuste', color='red')
plt.xlim(2000, 35000)
#plt.ylim(0, 30)
#plt.xscale('log')
plt.xlabel('w [rad/s]')
plt.ylabel('Phase [\u00b0]')
plt.legend()
plt.grid()
plt.show()

print("Parâmetros Ajustados para o ajuste à fase:")
print("K:", A2_params_phase[0])
print("w0:", A2_params_phase[1])
print("B:", A2_params_phase[2])

print(10**(T_dB(A2_params_phase[1], A2_params_phase[0], A2_params_phase[1], A2_params_phase[2])/20))

