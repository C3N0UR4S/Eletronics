# imports
import math
import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import openpyxl  

A1 = 2.30

# Def K, Q, w0
C1, C2, R2, R4, R5, R6, R11 = 4.7e-9, 4.7e-9, 100e3, 10e3, 100e3, 10e3, 10e3

w0 = np.sqrt(R5/(R2*R4*R11*C1*C2))
Q = w0*C1*R6

def T1_dB(x, K, w0, Q):
    return 20 * np.log10(np.abs(-K*w0*1j*x / (w0**2 + 1j * x * w0 / Q - x**2)))

# Caminho relativo para o arquivo Excel dentro da pasta "lab_02"
caminho_arquivo_excel = 'Lab_02/1_6_4.xlsx'

# Carregando o arquivo Excel
workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
sheet1 = workbook['P2=5k']
sheet2 = workbook['P2=1k']
sheet3 = workbook['P2=10k']

T1_5k_f, T1_1k_f, T1_10k_f= [], [], [] 
T1_5k_A2, T1_1k_A2, T1_10k_A2 = [], [], [] 
T1_5k_phase, T1_1k_phase, T1_10k_phase = [], [], []  

for row in sheet1.iter_rows(min_row=3, values_only=True):
    f, A2, phase = row[0:3]
    print(row)
    T1_5k_f.append(float(f))
    T1_5k_A2.append(float(A2))
    T1_5k_phase.append(float(phase))

for row in sheet2.iter_rows(min_row=3, values_only=True):
    f, A2, phase = row[0:3]
    print(row)
    T1_1k_f.append(float(f))
    T1_1k_A2.append(float(A2))
    T1_1k_phase.append(float(phase))

for row in sheet3.iter_rows(min_row=3, values_only=True):
    f, A2, phase = row[0:3]
    print(row)
    T1_10k_f.append(float(f))
    T1_10k_A2.append(float(A2))
    T1_10k_phase.append(float(phase))

T1_5k_w, T1_1k_w, T1_10k_w = [f*2*math.pi*10**3 for f in T1_5k_f], [f*2*np.pi*10**3 for f in T1_1k_f], [f*2*np.pi*10**3 for f in T1_10k_f]
T1_5k_A2_dB, T1_1k_A2_dB, T1_10k_A2_dB = [20*np.log10(np.abs(A2/A1)) for A2 in T1_5k_A2], [20*np.log10(np.abs(A2/A1)) for A2 in T1_1k_A2] , [20*np.log10(np.abs(A2/A1)) for A2 in T1_10k_A2]

#  curvas teóricas
w_values = np.linspace(0, max(T1_5k_w), 10000)
P2_5k = 5e3
K_5k = 1/(w0*P2_5k*C1)
T1_5k_magnitude_dB = T1_dB(w_values, K_5k, w0, Q)

P2_1k = 1e3
K_1k = 1/(w0*P2_1k*C1)
T1_1k_magnitude_dB = T1_dB(w_values, K_1k, w0, Q)

P2_10k = 10e3
K_10k = 1/(w0*P2_10k*C1)
T1_10k_magnitude_dB = T1_dB(w_values, K_10k, w0, Q)

T1_5k_initial_guess = (K_5k, w0, 1.0)  # (K, w0, Q)
T1_1k_initial_guess = (K_1k, w0, 1.0)  # (K, w0, Q)
T1_10k_initial_guess = (K_10k, w0, 1.0)  # (K, w0, Q)

T1_5k_params, T1_5k_covariance = curve_fit(T1_dB, T1_5k_w, T1_5k_A2_dB, p0=T1_5k_initial_guess)
T1_5k_x_fit = np.linspace(min(T1_5k_w), max(T1_5k_w), 1000)
T1_5k_y_fit = T1_dB(T1_5k_x_fit, *T1_5k_params)

T1_1k_params, T1_1k_covariance = curve_fit(T1_dB, T1_1k_w, T1_1k_A2_dB, p0=T1_1k_initial_guess)
T1_1k_x_fit = np.linspace(min(T1_1k_w), max(T1_1k_w), 1000)
T1_1k_y_fit = T1_dB(T1_1k_x_fit, *T1_1k_params)

T1_10k_params, T1_10k_covariance = curve_fit(T1_dB, T1_10k_w, T1_10k_A2_dB, p0=T1_10k_initial_guess)
T1_10k_x_fit = np.linspace(min(T1_10k_w), max(T1_10k_w), 1000)
T1_10k_y_fit = T1_dB(T1_10k_x_fit, *T1_10k_params)

plt.scatter(T1_1k_w, T1_1k_A2_dB, label = "Pontos experimentais P2=1k", color='black')
plt.plot(w_values, T1_1k_magnitude_dB, label='Curva Teórica P2=1k', color='purple')
plt.plot(T1_1k_x_fit, np.real(T1_1k_y_fit), 'g-', label='Ajuste P2=1k', color = 'yellow')

plt.scatter(T1_5k_w, T1_5k_A2_dB, label = "Pontos experimentais P2=5k")
plt.plot(w_values, T1_5k_magnitude_dB, label='Curva Teórica P2=5k', color='green')
plt.plot(T1_5k_x_fit, np.real(T1_5k_y_fit), 'g-', label='Ajuste P2=5k', color='red')

plt.scatter(T1_10k_w, T1_10k_A2_dB, label = "Pontos experimentais P2=10k", color='olivedrab')
plt.plot(w_values, T1_10k_magnitude_dB, label='Curva Teórica P2=10k', color='orange')
plt.plot(T1_10k_x_fit, np.real(T1_10k_y_fit), 'g-', label='Ajuste P2=10k', color = 'deeppink')

plt.xlim(0, 140000)
plt.ylim(-20, 25)
plt.xlabel('w [rad/s]')
plt.ylabel('|T| [dB]')
plt.legend()
plt.grid()
plt.savefig('T1_P2_gain.pdf')
plt.show()

print("Parâmetros Ajustados para T1 com P2 = 1k:")
print("K:", T1_1k_params[0])
print("w0:", T1_1k_params[1])
print("Q:", T1_1k_params[2])

print("Parâmetros Ajustados para T1 com P2 = 5k:")
print("K:", T1_5k_params[0])
print("w0:", T1_5k_params[1])
print("Q:", T1_5k_params[2])

print("Parâmetros Ajustados para T1 com P2 = 10k:")
print("K:", T1_10k_params[0])
print("w0:", T1_10k_params[1])
print("Q:", T1_10k_params[2])


#T1_10k_phase_degrees = np.degrees(np.angle(-K_10k*w0*1j*w_values / (w0**2 + 1j * w_values * w0 / Q - w_values**2)))
#T1_5k_phase_degrees = np.degrees(np.angle(-K_5k*w0*1j*w_values / (w0**2 + 1j * w_values * w0 / Q - w_values**2)))
#T1_1k_phase_degrees = np.degrees(np.angle(-K_1k*w0*1j*w_values / (w0**2 + 1j * w_values * w0 / Q - w_values**2)))
#
#plt.scatter(T1_1k_w, T1_1k_phase, label = "Pontos experimentais P2=1k", color = 'black')
#plt.plot(w_values, T1_1k_phase_degrees, label = 'Curva Teórica P2=1k', color = 'green')
#
#plt.scatter(T1_5k_w, T1_5k_phase, label = "Pontos experimentais P2=5k", color = 'red')
#plt.plot(w_values, T1_5k_phase_degrees, label = 'Curva Teórica P2=5k', color = 'blue')
#
#plt.scatter(T1_10k_w, T1_10k_phase, label = "Pontos experimentais P2=10k", color = 'orange')
#plt.plot(w_values, T1_10k_phase_degrees, label = 'Curva Teórica P2=10k', color = 'yellow')
#
#plt.xlabel('w [rad/s]')
#plt.ylabel('Phase [\u00b0]')
##plt.xlim(0000, 65000)
##plt.ylim(-15, 5)
##plt.title('Filtro highpass - ganho em dB')
#plt.legend()
#plt.grid(True)
##plt.savefig('T1_P2_phase.pdf')
#plt.show()

k = [T1_1k_params[0], T1_5k_params[0], T1_10k_params[0]]
P2 = [1e3, 5e3, 10e3]

#  curva teórica
P2_values = np.linspace(0, 10e3, 10000)

def k_P2(P2, C1, w0):
    return 1 / (P2 * C1 * w0)

k_values = k_P2(P2_values, C1, w0)

# Defina a função de ajuste
def func(x, a):
    return a / x

# Realize o ajuste de curva
params, covariance = curve_fit(func, P2, k)

# Parâmetros ajustados
a_fit = params[0]

# Gere pontos da curva ajustada
x_fit = np.linspace(min(P2), max(P2), 100)
y_fit = func(x_fit, a_fit)

plt.scatter(P2, k , label = "Pontos experimentais", color = 'black')
plt.plot(P2_values, k_values, label='Curva Teórica', color='orange')
plt.plot(x_fit, y_fit, label='Ajuste', color='red')
plt.xlabel('P2 [Ω]')
plt.ylabel('K')
plt.legend()
plt.show()

print(f"Valor ajustado para a: {a_fit}")
print(f"Valor teórico para a: {np.sqrt(R2*R4*R11*C2/R5/C1)}")

