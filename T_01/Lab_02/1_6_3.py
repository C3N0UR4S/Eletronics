# imports
import math
import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import openpyxl  

A1 = 2.30

# Def K, Q, w0
C1, C2, R2, R4, R5, R6, R11, P2 = 4.7e-9, 4.7e-9, 100e3, 10e3, 100e3, 10e3, 10e3, 10e3

w0 = np.sqrt(R5/(R2*R4*R11*C1*C2))
K = 1/(w0*P2*C1)
Q = w0*C1*R6

def T1_dB(x, K, w0, Q):
    return 20 * np.log10(np.abs(-K*w0*1j*x / (w0**2 + 1j * x * w0 / Q - x**2)))

def T2_dB(x, K, w0, Q):
    return 20 * np.log10(np.abs(K*w0**2 / (w0**2 + 1j * x * w0 / Q - x**2)))

def T3_dB(x, K, w0, Q):
    return 20 * np.log10(np.abs(-K*w0**2 / (w0**2 + 1j * x * w0 / Q - x**2)))


# Caminho relativo para o arquivo Excel dentro da pasta "lab_02"
caminho_arquivo_excel = 'Lab_02/1_6_3.xlsx'

# Carregando o arquivo Excel
workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
sheet1 = workbook['T1']
sheet2 = workbook['T2']
sheet3 = workbook['T3']

T1_f, T2_f, T3_f = [], [], []
T1_A2, T2_A2, T3_A2 = [], [], []
T1_phase, T2_phase, T3_phase = [], [], []

for row in sheet1.iter_rows(min_row=2, values_only=True):
    f, A2, phase = row[0:3]
    print(row)
    T1_f.append(float(f))
    T1_A2.append(float(A2))
    T1_phase.append(float(phase))

for row in sheet2.iter_rows(min_row=2, values_only=True):
    f, A2, phase = row[0:3]
    print(row)
    T2_f.append(float(f))
    T2_A2.append(float(A2))
    T2_phase.append(float(phase))

for row in sheet3.iter_rows(min_row=2, values_only=True):
    f, A2, phase = row[0:3]
    print(row)
    T3_f.append(float(f))
    T3_A2.append(float(A2))
    T3_phase.append(float(phase))


T1_w, T2_w, T3_w = [f*2*math.pi*10**3 for f in T1_f], [f*2*np.pi*10**3 for f in T2_f], [f*2*np.pi*10**3 for f in T3_f]
T1_A2_dB, T2_A2_dB, T3_A2_dB = [20*np.log10(np.abs(A2/A1)) for A2 in T1_A2], [20*np.log10(np.abs(A2/A1)) for A2 in T2_A2],[20*np.log10(np.abs(A2/A1)) for A2 in T3_A2]

#  curvas teóricas
w_values = np.linspace(0, max(T1_w), 10000)
T1_magnitude_dB = T1_dB(w_values, K, w0, Q)
T2_magnitude_dB = T2_dB(w_values, K, w0, Q)
T3_magnitude_dB = T3_dB(w_values, K, w0, Q)


T1_initial_guess = (1.0, w0, 1.0)  # (K, w0, Q)

T1_params, T1_covariance = curve_fit(T1_dB, T1_w, T1_A2_dB, p0=T1_initial_guess)
T1_x_fit = np.linspace(min(T1_w), max(T1_w), 1000)
T1_y_fit = T1_dB(T1_x_fit, *T1_params)

plt.scatter(T1_w, T1_A2_dB, label = "Pontos experimentais")
plt.plot(w_values, T1_magnitude_dB, label='Curva Teórica', color='green')
plt.plot(T1_x_fit, np.real(T1_y_fit), 'g-', label='Ajuste', color='red')
plt.xlim(0, 65000)
plt.ylim(-20, 2.5)
plt.xlabel('w [rad/s]')
plt.ylabel('|T| [dB]')
plt.legend()
plt.grid()
plt.savefig('TT_T1_gain.pdf')
plt.show()

T1_phase_degrees = np.degrees(np.angle(-K*w0*1j*w_values / (w0**2 + 1j * w_values * w0 / Q - w_values**2)))

plt.scatter(T1_w, T1_phase, label = "Pontos experimentais")
plt.plot(w_values, T1_phase_degrees, label = 'Curva Teórica', color = 'green')
plt.xlabel('w [rad/s]')
plt.ylabel('Phase [\u00b0]')
plt.xlim(0000, 65000)
#plt.ylim(-15, 5)
#plt.title('Filtro highpass - ganho em dB')
plt.legend()
plt.grid(True)
plt.savefig('TT_T1_phase.pdf')
plt.show()

print("Parâmetros Ajustados para T1:")
print("K:", T1_params[0])
print("w0:", T1_params[1])
print("Q:", T1_params[2])

T2_initial_guess = (1.0, w0, 1.0)  # (K, w0, Q)

T2_params, T2_covariance = curve_fit(T2_dB, T2_w, T2_A2_dB, p0=T2_initial_guess)
T2_x_fit = np.linspace(min(T2_w), max(T2_w), 1000)
T2_y_fit = T2_dB(T2_x_fit, *T2_params)

plt.scatter(T2_w, T2_A2_dB, label = "Pontos experimentais", color = 'brown')
plt.plot(w_values, T2_magnitude_dB, label='Curva Teórica')
plt.plot(T2_x_fit, np.real(T2_y_fit), 'g-', label='Ajuste', color='green')
plt.xlim(0, 65000)
plt.ylim(-20, 2.5)
plt.xlabel('w [rad/s]')
plt.ylabel('|T| [dB]')
plt.legend()
plt.grid()
plt.savefig('TT_T2_gain.pdf')
plt.show()

T2_phase_degrees = np.degrees(np.angle(K*w0**2 / (w0**2 + 1j * w_values * w0 / Q - w_values**2)))

plt.scatter(T2_w, T2_phase, label = "Pontos experimentais")
plt.plot(w_values, T2_phase_degrees, label = 'Curva Teórica', color = 'green')
plt.xlabel('w [rad/s]')
plt.ylabel('Phase [\u00b0]')
plt.xlim(0000, 65000)
#plt.ylim(-15, 5)
#plt.title('Filtro highpass - ganho em dB')
plt.legend()
plt.grid(True)
plt.savefig('TT_T2_phase.pdf')
plt.show()

print("Parâmetros Ajustados para T2:")
print("K:", T2_params[0])
print("w0:", T2_params[1])
print("Q:", T2_params[2])

T3_initial_guess = (1.0, w0, 1.0)  # (K, w0, Q)

T3_params, T3_covariance = curve_fit(T3_dB, T3_w, T3_A2_dB, p0=T3_initial_guess)
T3_x_fit = np.linspace(min(T3_w), max(T3_w), 1000)
T3_y_fit = T3_dB(T3_x_fit, *T3_params)

plt.scatter(T3_w, T3_A2_dB, label = "Pontos experimentais")
plt.plot(w_values, T3_magnitude_dB, label='Curva Teórica', color='red')
plt.plot(T3_x_fit, np.real(T3_y_fit), 'g-', label='Ajuste', color='green')
plt.xlim(0, 65000)
plt.ylim(-20, 2.5)
plt.xlabel('w [rad/s]')
plt.ylabel('|T| [dB]')
plt.legend()
plt.grid()
plt.savefig('TT_T3_gain.pdf')
plt.show()

T3_phase_degrees = np.degrees(np.angle(-K*w0**2 / (w0**2 + 1j * w_values * w0 / Q - w_values**2)))

plt.scatter(T3_w, T3_phase, label = "Pontos experimentais")
plt.plot(w_values, T3_phase_degrees, label = 'Curva Teórica', color = 'green')
plt.xlabel('w [rad/s]')
plt.ylabel('Phase [\u00b0]')
plt.xlim(0000, 65000)
#plt.ylim(-15, 5)
#plt.title('Filtro highpass - ganho em dB')
plt.legend()
plt.grid(True)
plt.savefig('TT_T3_phase.pdf')
plt.show()

print("Parâmetros Ajustados para T3:")
print("K:", T3_params[0])
print("w0:", T3_params[1])
print("Q:", T3_params[2])